from pathlib import Path

from openpyxl import Workbook, load_workbook


DATA_DIR = Path("data")
DATA_DIR.mkdir(exist_ok=True)

EXCEL_FILE = DATA_DIR / "avito.xlsx"


class ExcelLogger:

    def __init__(self):

        if not EXCEL_FILE.exists():

            wb = Workbook()

            ws = wb.active

            ws.append([
                "ID",
                "Название",
                "Цена",
                "Город",
                "Ссылка",
                "Дата"
            ])

            wb.save(EXCEL_FILE)

    def add_item(self, item):

        wb = load_workbook(EXCEL_FILE)

        ws = wb.active

        ws.append([
            item["id"],
            item["title"],
            item["price"],
            item["city"],
            item["url"],
            item["found_at"]
        ])

        wb.save(EXCEL_FILE)