from pathlib import Path
from openpyxl import Workbook, load_workbook

DATA_DIR = Path("data")
DATA_DIR.mkdir(exist_ok=True)

FILE = DATA_DIR / "price_history.xlsx"


class PriceLogger:

    def __init__(self):

        if not FILE.exists():

            wb = Workbook()

            ws = wb.active

            ws.append([
                "Дата",
                "Название",
                "Было",
                "Стало",
                "Разница",
                "Ссылка"
            ])

            wb.save(FILE)

    def add(self, item, old_price):

        wb = load_workbook(FILE)

        ws = wb.active

        ws.append([
            item["found_at"],
            item["title"],
            old_price,
            item["price"],
            item["price"] - old_price,
            item["url"]
        ])

        wb.save(FILE)